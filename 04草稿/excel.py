import openpyxl
import csv
import codecs

work_book = openpyxl.Workbook()

# 创建sheet
work_sheet = work_book.create_sheet(title=u"cpu和内存")

# 打开csv文件，没有异常捕获，没有定义成函数,做成模块更好
csvfile = open(r'D:\aaaa.csv', 'rt', encoding="utf-8")

# 获取csv.reader
lines = csv.reader(csvfile)

# # row
row = 0
#
# # sheet第一行，就叫他标题吧
# title_list = [u'采集时间', u'主机', u'CPU使用率（%）', u'MEM使用率（%）']
#
# # 写入第一行，标题
# for i in range(1, title_len + 1):
#     work_sheet.cell(row=row, column=i).value = title_list[i - 1]

# 从第二行开始写入从csv读取的内容
for line in lines:
    lin = 1
    for i in line:
        work_sheet.cell(row=row + 1, column=lin).value = i
        lin += 1
    row += 1

# 关闭文件
csvfile.close()

# 保存工作表
work_book.save(r'D:\ecmp.xlsx')
